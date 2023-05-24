from openpyxl import load_workbook
import pandas as pd

def write_sheet_to_excel(df, sheet_name, input_file_path):
    # Load the existing Excel file
    book = load_workbook(input_file_path)

    # If the sheet name already exists, remove it
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        book.remove(sheet)

    # Save the changes to the existing Excel file
    book.save(input_file_path)
    book.close()

    # Write the new sheet to the Excel file
    with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)